# -*- coding: utf-8 -*-

import os
import shutil
import time
import re
import openpyxl
from ast import literal_eval
from PIL import Image
from PIL import ImageStat
from owslib.wms import WebMapService

czas1 = time.time()
#definicja klasy powiatu, obiekty zawierają nazwe, teryt, adres wms, wspolrzedne
class powiat(object):
    def __init__(self, nazwa, teryt, adres, szerokosc, dlugosc, X, Y, szer_p_losowego, dlug_p_losowego, X_p_losowego, Y_p_losowego, nazwa_warstwy, stan = "Dziala" ):
        self.nazwa = nazwa
        self.teryt = teryt
        self.adres = adres
        self.szerokosc = szerokosc
        self.dlugosc = dlugosc
        self.X = X
        self.Y = Y
        self.szer_p_losowego = szer_p_losowego
        self.dlug_p_losowego = dlug_p_losowego
        self.X_p_losowego = X_p_losowego
        self.Y_p_losowego = Y_p_losowego
        self.stan = stan
        self.nazwa_warstwy = nazwa_warstwy


def ZapiszObraz(obiekt_powiatu, wms, warstwa, size = (1200,800), format = 'image/png', transparent = True):

    wms.getOperationByName('GetMap').methods[0]['url'] = obiekt_powiatu.adres

    sciezka = 'D:\\Piotr\\Magisterka\\RISE\\obrazy' + '\\' + obiekt_powiatu.nazwa + '.png'

    lista_bledow = []
    czas_start = time.time()
    try:
        img = wms.getmap(   layers = [warstwa],
                                    srs ='EPSG:2180',
                                    bbox = bb92,
                                    size = size,
                                    format = format,
                                    transparent = True,
                                    styles = [''])
        out = open(sciezka, 'wb')
        out.write(img.read())
        out.close()

        czas_koniec = time.time()
        czas = czas_koniec - czas_start

        Raport(lista_bledow, czas)
    except Exception as blad:
        lista_bledow.append(4)
        Raport(lista_bledow)



def Raport(item, data):

    for wiersz in range(1, arkusz.max_row + 1):
        wiersz_excel = str(arkusz.cell(row = wiersz, column = 2).value)
        teryt = os.path.splitext(item)[0]
        if wiersz_excel == teryt:
            wiersz = wiersz
            break


    arkusz.cell(row = wiersz, column = 17).value = data






plik_z_terytami = open('D:\\Piotr\\Magisterka\\RISE\\teryty.txt', 'r')

#otworzenie excela,
excel = openpyxl.load_workbook('raport_kiip.xlsx')
arkusz = excel.get_sheet_by_name('parametry')

kolumna_powiatu = arkusz.max_column

#otworzenie pliku z powiatami
plik_z_powiatami = open('D:\\Piotr\\Magisterka\\RISE\\lista_wms_kiip.txt','r')

os.chdir('D:\\Piotr\\Magisterka\\RISE\\pliki_xml\\')
lista_plikow = os.listdir('D:\\Piotr\\Magisterka\\RISE\\pliki_xml\\')

for item in lista_plikow:
    plik = open(item, 'r')
    tekst = plik.read()

    z1 = tekst.find('DateOfLastRevision>')
    z2 = tekst.find('</ns2:DateOfLastRevision>')

    dlugosc = len('DateOfLastRevision>')
    if z1 > 0:
        data = tekst[z1 + dlugosc:z2]
    else:
        data = 'brak'

    Raport(item, data)







excel.save('raport_kiip.xlsx')
excel.close()

czas2 = time.time()
czas_sekundy = czas2 - czas1
czas_minuty = czas_sekundy / 60

print '\n'
print 'CZAS CALKOWITY (min): ' + str(round(czas_minuty,2))
print '\n' + "KONIEC"

