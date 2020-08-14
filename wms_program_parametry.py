# -*- coding: utf-8 -*-

import os
import shutil
import time
import re
import openpyxl
from ast import literal_eval
from PIL import Image
from PIL import ImageStat
from owslib.wms import WebMapService

czas1 = time.time()
#definicja klasy powiatu, obiekty zawierają nazwe, teryt, adres wms, wspolrzedne
class powiat(object):
    def __init__(self, nazwa, teryt, adres, szerokosc, dlugosc, X, Y, szer_p_losowego, dlug_p_losowego, X_p_losowego, Y_p_losowego, nazwa_warstwy, stan = "Dziala" ):
        self.nazwa = nazwa
        self.teryt = teryt
        self.adres = adres
        self.szerokosc = szerokosc
        self.dlugosc = dlugosc
        self.X = X
        self.Y = Y
        self.szer_p_losowego = szer_p_losowego
        self.dlug_p_losowego = dlug_p_losowego
        self.X_p_losowego = X_p_losowego
        self.Y_p_losowego = Y_p_losowego
        self.stan = stan
        self.nazwa_warstwy = nazwa_warstwy


def ZapiszObraz(obiekt_powiatu, wms, warstwa, size = (1200,800), format = 'image/png', transparent = True):

    wms.getOperationByName('GetMap').methods[0]['url'] = obiekt_powiatu.adres

    sciezka = 'D:\\Piotr\\Magisterka\\RISE\\obrazy' + '\\' + obiekt_powiatu.nazwa + '.png'

    lista_bledow = []
    czas_start = time.time()
    try:
        img = wms.getmap(   layers = [warstwa],
                                    srs ='EPSG:2180',
                                    bbox = bb92,
                                    size = size,
                                    format = format,
                                    transparent = True,
                                    styles = [''])
        out = open(sciezka, 'wb')
        out.write(img.read())
        out.close()

        czas_koniec = time.time()
        czas = czas_koniec - czas_start

        Raport(lista_bledow, czas)
    except Exception as blad:
        lista_bledow.append(4)
        Raport(lista_bledow)



def Raport(lista_parametrow):

    for wiersz in range(1, arkusz.max_row + 1):
        wiersz_excel = str(arkusz.cell(row = wiersz, column = 2).value)
        teryt = obiekt_powiatu.teryt
        if wiersz_excel == teryt:
            wiersz = wiersz
            break

    a = 0
    for item in lista_parametrow:

        if isinstance(item, str) == True:
            lista_parametrow[a] = item.decode('windows-1250')
        elif isinstance(item, list) == True:
            lista_parametrow[a] = str(item)
        a += 1


    arkusz.cell(row = wiersz, column = 3).value = lista_parametrow[0]
    arkusz.cell(row = wiersz, column = 4).value = lista_parametrow[1]
    arkusz.cell(row = wiersz, column = 5).value = lista_parametrow[2]
    arkusz.cell(row = wiersz, column = 6).value = lista_parametrow[3]
    arkusz.cell(row = wiersz, column = 7).value = lista_parametrow[4]
    arkusz.cell(row = wiersz, column = 8).value = lista_parametrow[5]
    arkusz.cell(row = wiersz, column = 9).value = lista_parametrow[6]
    arkusz.cell(row = wiersz, column = 10).value = lista_parametrow[7]
    arkusz.cell(row = wiersz, column = 11).value = lista_parametrow[8]
    arkusz.cell(row = wiersz, column = 12).value = lista_parametrow[9]
    arkusz.cell(row = wiersz, column = 13).value = lista_parametrow[10]
    arkusz.cell(row = wiersz, column = 14).value = lista_parametrow[11]
    arkusz.cell(row = wiersz, column = 15).value = lista_parametrow[12]
    arkusz.cell(row = wiersz, column = 16).value = lista_parametrow[13]





plik_z_terytami = open('D:\\Piotr\\Magisterka\\RISE\\teryty.txt', 'r')

#otworzenie excela,
excel = openpyxl.load_workbook('raport_kiip.xlsx')
arkusz = excel.get_sheet_by_name('parametry')

kolumna_powiatu = arkusz.max_column

#otworzenie pliku z powiatami
plik_z_powiatami = open('D:\\Piotr\\Magisterka\\RISE\\test.txt','r')

# iterowanie po kolejnych liniach, tworzenie 'obiektu_powiatu'
i = 2
j = 0
for linia in plik_z_powiatami:

    lista = linia.split('\t')
    obiekt_powiatu = powiat(lista[0], lista[1], lista[2], lista[5], lista[6], lista[7], lista[8], lista[9], lista[10], lista[11], lista[12], lista[13].rstrip())


    print obiekt_powiatu.nazwa, str(obiekt_powiatu.teryt)
    i += 1

    try:
        wms = WebMapService(obiekt_powiatu.adres)
    except AttributeError as at_error:      # warunek na wersje 1.3.0 wmsa np. Białystok
        if str(at_error) == "'NoneType' object has no attribute 'find'":
            wms = WebMapService(obiekt_powiatu.adres, version = '1.3.0')
    except ValueError as va_error:          # warunek na brak zdefiniowanych styli np. powiat bielski
        from owslib.testpiotr.wms import WebMapService
        wms = WebMapService(obiekt_powiatu.adres)
    except:                                 # warunek gdy nie chce się połączyc
        obiekt_powiatu.stan = 'Nie dziala'

    #jeżeli wms odpowiada, pytam o warstwy i szukam takiej, ktora ma w nazwie słowo 'działki'
    if obiekt_powiatu.stan == 'Dziala':

        X1 = float(obiekt_powiatu.X_p_losowego) - 400
        Y1 = float(obiekt_powiatu.Y_p_losowego) - 600
        X2 = float(obiekt_powiatu.X_p_losowego) + 400
        Y2 = float(obiekt_powiatu.Y_p_losowego) + 600
        bb92 = (Y1, X1, Y2, X2)

        if obiekt_powiatu.teryt == "3261": # warunek na Koszalin, ogranicznie rozmiarow
            X1 = float(obiekt_powiatu.X_p_losowego) - 100
            Y1 = float(obiekt_powiatu.Y_p_losowego) - 200
            X2 = float(obiekt_powiatu.X_p_losowego) + 100
            Y2 = float(obiekt_powiatu.Y_p_losowego) + 200
            bb92 = (Y1, X1, Y2, X2)


        xml = wms.getServiceXML()

        plik_xml = open('D:\\Piotr\\Magisterka\\RISE\\pliki_xml\\' + obiekt_powiatu.teryt + '.txt', 'wb')
        plik_xml.write(xml)
        plik_xml.close()



        tytul = wms.identification.title.encode('windows-1250')

        abstract = wms.identification.abstract

        slowa_kluczowe = wms.identification.keywords

        zawartosc = list(wms.contents)

        uklady = wms["dzialki"].crsOptions

        formaty = wms.getOperationByName('GetMap').formatOptions

        #kontakt
        nazwa = wms.provider.name
        adres_url = wms.provider.url
        try:
            ulica = wms.provider.contact.address
            miasto = wms.provider.contact.city
            panstwo = wms.provider.contact.country
            kod_pocztowy = wms.provider.contact.postcode
            nazwisko = wms.provider.contact.name
            email = wms.provider.contact.email
        except:
            ulica = "brak"
            miasto = "brak"
            panstwo = "brak"
            kod_pocztowy = "brak"
            nazwisko = "brak"
            email = "brak"
            pass

        obiekt_powiatu.nazwa_warstwy = obiekt_powiatu.nazwa_warstwy.decode('windows-1250')
        obiekt_powiatu.nazwa_warstwy = obiekt_powiatu.nazwa_warstwy.encode('utf-8')

        lista_parametrow = [tytul, abstract, slowa_kluczowe, zawartosc, uklady, formaty, nazwa, adres_url, ulica, miasto, panstwo, kod_pocztowy, nazwisko, email]
        Raport(lista_parametrow)


excel.save('raport_kiip.xlsx')
excel.close()

czas2 = time.time()
czas_sekundy = czas2 - czas1
czas_minuty = czas_sekundy / 60

print '\n'
print 'CZAS CALKOWITY (min): ' + str(round(czas_minuty,2))
print '\n' + "KONIEC"

