# -*- coding: utf-8 -*-

import os
import shutil
import glob
import time
import re
import openpyxl
from ast import literal_eval
from PIL import Image
from PIL import ImageStat
from PIL import ImageColor
from owslib.wms import WebMapService

czas1 = time.time()
#definicja klasy powiatu, obiekty zawierają nazwe, teryt, adres wms, wspolrzedne
class powiat(object):
    def __init__(self, nazwa, teryt, adres, szerokosc, dlugosc, X, Y, szer_p_losowego, dlug_p_losowego, X_p_losowego, Y_p_losowego, nazwa_warstwy, stan = "Dziala" ):
        self.nazwa = nazwa
        self.teryt = teryt
        self.adres = adres
        self.szerokosc = szerokosc
        self.dlugosc = dlugosc
        self.X = X
        self.Y = Y
        self.szer_p_losowego = szer_p_losowego
        self.dlug_p_losowego = dlug_p_losowego
        self.X_p_losowego = X_p_losowego
        self.Y_p_losowego = Y_p_losowego
        self.stan = stan
        self.nazwa_warstwy = nazwa_warstwy

def SprawdzObraz(obiekt_powiatu):

    lista_obrazow = glob.glob('D:\\Piotr\\Magisterka\\RISE\\obrazy_100x300x300\\' + obiekt_powiatu.teryt + '\\' + '*.png')

    lista_stanow_obrazow = []
    for item in lista_obrazow:
        obraz = Image.open(item)
        obraz = obraz.convert('RGBA')

        liczba_pikseli = obraz.width * obraz.height

        histogram = obraz.histogram()
        r = histogram[0:256]
        g = histogram[256:512]
        b = histogram[512:768]
        a = histogram[768:1024]

        print r, '\n'
        print g, '\n'
        print b, '\n'
        print a, '\n'

        a = os.path.split(item)
        nazwa_obrazu = os.path.splitext(a[1])[0]

        #warunek na calkowicie puste obrazy
        if (r[0],g[0],b[0]) == (liczba_pikseli,liczba_pikseli,liczba_pikseli) or (r[255],g[255],b[255]) == (liczba_pikseli,liczba_pikseli,liczba_pikseli):
            lista_stanow_obrazow.append(1)
        else:
            #okreslenie ktore piksele znajduja sie w tle (biale i nieprzezroczyste czy czarne i całkowicie przezroczyste)
            if histogram[0] > histogram[255]:
                tlo = 0
            else:
                tlo = 255

            #wybieram najmniejsza z najwiekszych liczb w kazdej liscie r,g,b.
            lista_x = [r[tlo], g[tlo], b[tlo]]
            liczba_pikseli_tla = min(lista_x)

            liczba_pikseli_dzialek = liczba_pikseli - liczba_pikseli_tla

            if tlo == 0:
                r_max = max(r[1:])
                g_max = max(g[1:])
                b_max = max(b[1:])

                r_max = r.index(r_max)
                g_max = g.index(g_max)
                b_max = b.index(b_max)
            elif tlo == 255:
                r_max = max(r[:255])
                g_max = max(g[:255])
                b_max = max(b[:255])

                if r_max == 0:
                    r_max = r[255]
                elif g_max == 0:
                    g_max = g[255]
                elif b_max == 0:
                    b_max = b[255]

                r_max = r.index(r_max)
                g_max = g.index(g_max)
                b_max = b.index(b_max)

            if r_max in range(58,70) and g_max in range(154,166) and b_max in range(249,261):
                lista_stanow_obrazow.append(2)  # 64, 160, 255
            else:
                lista_stanow_obrazow.append(3)

    Raport(lista_stanow_obrazow)

    """
    obraz_statystyka = ImageStat.Stat(histogram)
    srednie = obraz_statystyka.mean

    srednia_r_dzialek = (srednie[0] * liczba_pikseli - liczba_pikseli_tla * tlo)/ (liczba_pikseli_dzialek)
    srednia_g_dzialek = (srednie[1] * liczba_pikseli - liczba_pikseli_tla * tlo)/ (liczba_pikseli_dzialek)
    srednia_b_dzialek = (srednie[2] * liczba_pikseli - liczba_pikseli_tla * tlo)/ (liczba_pikseli_dzialek)



    if (round(srednia_r_dzialek) in range(58,70)) and (round(srednia_g_dzialek) in range(154,166)) and (round(srednia_b_dzialek) in range(249,261)):
        lista_stanow_obrazow.append(2)  # 64, 160, 255
    else:
        lista_stanow_obrazow.append(3)

    Raport(lista_stanow_obrazow)

    print nazwa_obrazu, srednia_r_dzialek, srednia_g_dzialek, srednia_b_dzialek, '\n'
    """
def ZapiszObraz(obiekt_powiatu, wms, warstwa, size = (1200,800), format = 'image/png', transparent = True):

    wms.getOperationByName('GetMap').methods[0]['url'] = obiekt_powiatu.adres

    sciezka = 'D:\\Piotr\\Magisterka\\RISE\\obrazy' + '\\' + obiekt_powiatu.nazwa + '.png'

    lista_bledow = []
    czas_start = time.time()
    try:
        img = wms.getmap(   layers = [warstwa],
                                    srs ='EPSG:2180',
                                    bbox = bb92,
                                    size = size,
                                    format = format,
                                    transparent = True,
                                    styles = [''])
        out = open(sciezka, 'wb')
        out.write(img.read())
        out.close()

        czas_koniec = time.time()
        czas = czas_koniec - czas_start

        Raport(lista_bledow, czas)
    except Exception as blad:
        lista_bledow.append(4)
        Raport(lista_bledow)

def Raport(lista_stanow_obrazow = []):

    puste = lista_stanow_obrazow.count(1)
    zgodne = lista_stanow_obrazow.count(2)
    niezgodne = lista_stanow_obrazow.count(3)

    for wiersz in range(1, arkusz.max_row + 1):
        wiersz_excel = str(arkusz.cell(row = wiersz, column = 2).value)
        teryt = obiekt_powiatu.teryt
        if wiersz_excel == teryt:
            wiersz = wiersz
            break

    arkusz.cell(row = wiersz, column = kolumna_powiatu).value = puste
    arkusz.cell(row = wiersz, column = kolumna_powiatu + 1).value = zgodne
    arkusz.cell(row = wiersz, column = kolumna_powiatu + 2).value = niezgodne

def ZapiszObraz100(obiekt_powiatu, wms, warstwa, liczba_obrazow = 50, size = (1200,800), format = 'image/png', transparent = True, bgcolor = '#FFFFFF'):

    wms.getOperationByName('GetMap').methods[0]['url'] = obiekt_powiatu.adres

    plik_z_punktami = open('D:\\Piotr\\Magisterka\\RISE\\punkty_losowe\\' + obiekt_powiatu.teryt + '\\' + obiekt_powiatu.teryt + '.txt', 'r')

    sciezka_obrazu = 'D:\\Piotr\\Magisterka\\RISE\\punkty_losowe\\' + obiekt_powiatu.teryt + '\\'

    lista_bledow = []
    k = 0
    czas_start = time.time()

    for linia in plik_z_punktami:
        k += 1
        if k > liczba_obrazow:
            break
        punkt = linia.split('\t')
        X1 = float(punkt[3]) - 400
        Y1 = float(punkt[2]) - 600
        X2 = float(punkt[3]) + 400
        Y2 = float(punkt[2]) + 600
        bb92 = (Y1, X1, Y2, X2)

        if obiekt_powiatu.teryt == "3261": # warunek na Koszalin, ogranicznie rozmiarow
            X1 = float(punkt[3]) - 100
            Y1 = float(punkt[2]) - 200
            X2 = float(punkt[3]) + 100
            Y2 = float(punkt[2]) + 200
            bb92 = (Y1, X1, Y2, X2)


        try:
            img = wms.getmap(   layers = [warstwa],
                                        srs ='EPSG:2180',
                                        bbox = bb92,
                                        size = size,
                                        format = format,
                                        transparent = True,
                                        bgcolor = bgcolor,
                                        styles = [''])
            out = open(sciezka_obrazu + str(k) + '.png', 'wb')
            out.write(img.read())
            out.close()

            if k % 10 == 0:
                print str(k)
        except Exception as blad:
            lista_bledow.append(1)
            if (k == 5) and (lista_bledow.count(1) == 5):
                lista_bledow.append(2)
                break



    czas_koniec = time.time()
    czas = czas_koniec - czas_start


    Raport(lista_bledow, czas)




#otworzenie excela,
excel = openpyxl.load_workbook('raport_kiip.xlsx')
arkusz = excel.get_sheet_by_name('obrazy')
data = time.strftime("%Y%m%d")
data = str(data[6:]) + '-' + str(data[4:6]) + '-' + str(data[0:4])

godzina = time.strftime("%H%M")
godzina = str(godzina[0:2]) + ':' + str(godzina[2:])

arkusz.cell(row = 1, column = arkusz.max_column + 1).value = data
arkusz.cell(row = 2, column = arkusz.max_column).value = godzina

kolumna_powiatu = arkusz.max_column

#otworzenie pliku z powiatami
plik_z_powiatami = open('D:\\Piotr\\Magisterka\\RISE\\lista_wms_kiip.txt','r')

# iterowanie po kolejnych liniach, tworzenie 'obiektu_powiatu'
i = 2
j = 0
for linia in plik_z_powiatami:

    lista = linia.split('\t')
    obiekt_powiatu = powiat(lista[0], lista[1], lista[2], lista[5], lista[6], lista[7], lista[8], lista[9], lista[10], lista[11], lista[12], lista[13].rstrip())


    print obiekt_powiatu.nazwa, str(obiekt_powiatu.teryt)

    SprawdzObraz(obiekt_powiatu)


excel.save('raport_kiip.xlsx')
excel.close()

czas2 = time.time()
czas_sekundy = czas2 - czas1
czas_minuty = czas_sekundy / 60

print '\n'
print 'CZAS CALKOWITY (min): ' + str(round(czas_minuty,2))
print '\n' + "KONIEC"

