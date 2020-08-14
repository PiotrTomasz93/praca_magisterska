# -*- coding: utf-8 -*-

import os
import shutil
import time
import re
import openpyxl
from ast import literal_eval
from PIL import Image
from PIL import ImageStat
from owslib.wms import WebMapService

czas1 = time.time()
#definicja klasy powiatu, obiekty zawierają nazwe, teryt, adres wms, wspolrzedne
class powiat(object):
    def __init__(self, nazwa, teryt, adres, szerokosc, dlugosc, X, Y, szer_p_losowego, dlug_p_losowego, X_p_losowego, Y_p_losowego, nazwa_warstwy, stan = "Dziala" ):
        self.nazwa = nazwa
        self.teryt = teryt
        self.adres = adres
        self.szerokosc = szerokosc
        self.dlugosc = dlugosc
        self.X = X
        self.Y = Y
        self.szer_p_losowego = szer_p_losowego
        self.dlug_p_losowego = dlug_p_losowego
        self.X_p_losowego = X_p_losowego
        self.Y_p_losowego = Y_p_losowego
        self.stan = stan
        self.nazwa_warstwy = nazwa_warstwy

def ZapiszObraz(obiekt_powiatu, wms, warstwa, size = (1200,800), format = 'image/png', transparent = True):

    wms.getOperationByName('GetMap').methods[0]['url'] = obiekt_powiatu.adres

    sciezka = 'D:\\Piotr\\Magisterka\\RISE\\obrazy' + '\\' + obiekt_powiatu.nazwa + '.png'

    lista_bledow = []
    czas_start = time.time()
    try:
        img = wms.getmap(   layers = [warstwa],
                                    srs ='EPSG:2180',
                                    bbox = bb92,
                                    size = size,
                                    format = format,
                                    transparent = True,
                                    styles = [''])
        out = open(sciezka, 'wb')
        out.write(img.read())
        out.close()

        czas_koniec = time.time()
        czas = czas_koniec - czas_start

        Raport(lista_bledow, czas)
    except Exception as blad:
        lista_bledow.append(4)
        Raport(lista_bledow)



def Raport(lista_bledow = [], czas = 0):
    czas = str(round(czas,2))

    bledy = lista_bledow.count(1)
    blad_2 = lista_bledow.count(2) #jezeli jest jeden blad o kodzie "2" tzn ze poczatkowe obrazy nie pobraly sie i wms nie dziala
    blad_3 = lista_bledow.count(3) #jezeli jest jeden blad o kodzie "3" tzn ze nie mozna sie polaczyc
    blad_4 = lista_bledow.count(4) #jezeli jest jeden blad o kodzie "4" tzn ze jest polaczenie ale nie mozna pobrac obrazu


    for wiersz in range(1, arkusz.max_row + 1):
        wiersz_excel = str(arkusz.cell(row = wiersz, column = 2).value)
        teryt = obiekt_powiatu.teryt
        if wiersz_excel == teryt:
            wiersz = wiersz
            break

    if blad_2 == 1:
        arkusz.cell(row = wiersz, column = kolumna_powiatu).value = "2"
        print "Liczba bledow: 5. Nie mozna pobrac obrazow."
        print 'Czas: ' + czas + ' sekund' + '\n'
    elif blad_3 == 1:
        arkusz.cell(row = wiersz, column = kolumna_powiatu).value = "3"
        print "Liczba bledow: 1. Brak polaczenia"
        print 'Czas: ' + czas + ' sekund' + '\n'
    elif blad_4 == 4:
        arkusz.cell(row = wiersz, column = kolumna_powiatu).value = "4"
        print "Liczba bledow: 1. Nie mozna pobrac obrazu."
        print 'Czas: ' + czas + ' sekund' + '\n'
    else:
        arkusz.cell(row = wiersz, column = kolumna_powiatu).value = czas
        #arkusz.cell(row = wiersz, column = kolumna_powiatu).value = czas + ' - ' + str(bledy)
        print "Liczba bledow: " + str(bledy)
        print 'Czas: ' + czas + ' sekund' + '\n'

def ZapiszObraz100(obiekt_powiatu, wms, warstwa, liczba_obrazow = 50, size = (1200,800), format = 'image/png', transparent = True, bgcolor = '#FFFFFF'):

    wms.getOperationByName('GetMap').methods[0]['url'] = obiekt_powiatu.adres

    plik_z_punktami = open('D:\\Piotr\\Magisterka\\RISE\\punkty_losowe\\' + obiekt_powiatu.teryt + '\\' + obiekt_powiatu.teryt + '.txt', 'r')

    sciezka_obrazu = 'D:\\Piotr\\Magisterka\\RISE\\punkty_losowe\\' + obiekt_powiatu.teryt + '\\'

    lista_bledow = []
    k = 0
    czas_start = time.time()

    for linia in plik_z_punktami:
        k += 1
        if k > liczba_obrazow:
            break
        punkt = linia.split('\t')
        X1 = float(punkt[3]) - 400
        Y1 = float(punkt[2]) - 600
        X2 = float(punkt[3]) + 400
        Y2 = float(punkt[2]) + 600
        bb92 = (Y1, X1, Y2, X2)

        if obiekt_powiatu.teryt == "3261": # warunek na Koszalin, ogranicznie rozmiarow
            X1 = float(punkt[3]) - 100
            Y1 = float(punkt[2]) - 200
            X2 = float(punkt[3]) + 100
            Y2 = float(punkt[2]) + 200
            bb92 = (Y1, X1, Y2, X2)


        try:
            img = wms.getmap(   layers = [warstwa],
                                        srs ='EPSG:2180',
                                        bbox = bb92,
                                        size = size,
                                        format = format,
                                        transparent = True,
                                        bgcolor = bgcolor,
                                        styles = [''])
            out = open(sciezka_obrazu + str(k) + '.png', 'wb')
            out.write(img.read())
            out.close()

            if k % 10 == 0:
                print str(k)
        except Exception as blad:
            lista_bledow.append(1)
            if (k == 5) and (lista_bledow.count(1) == 5):
                lista_bledow.append(2)
                break



    czas_koniec = time.time()
    czas = czas_koniec - czas_start


    Raport(lista_bledow, czas)


if os.path.exists('D:\\Piotr\\Magisterka\\RISE\\punkty_losowe') == True:
    shutil.rmtree('D:\\Piotr\\Magisterka\\RISE\\punkty_losowe', ignore_errors=True)

plik_z_terytami = open('D:\\Piotr\\Magisterka\\RISE\\teryty.txt', 'r')

os.makedirs('D:\\Piotr\\Magisterka\\RISE\\punkty_losowe')

for linia in plik_z_terytami:
    teryt = linia.rstrip()
    os.makedirs('D:\\Piotr\\Magisterka\\RISE\\punkty_losowe\\' + teryt)

    shutil.copy2('D:\\Piotr\\Magisterka\\punkty_losowe\\' + teryt + '\\' + teryt + '.txt', 'D:\\Piotr\\Magisterka\\RISE\\punkty_losowe\\' + teryt)



#otworzenie excela,
excel = openpyxl.load_workbook('raport_kiip.xlsx')
arkusz = excel.get_sheet_by_name('pojedyncze')
data = time.strftime("%Y%m%d")
data = str(data[6:]) + '-' + str(data[4:6]) + '-' + str(data[0:4])

godzina = time.strftime("%H%M")
godzina = str(godzina[0:2]) + ':' + str(godzina[2:])

arkusz.cell(row = 1, column = arkusz.max_column + 1).value = data
arkusz.cell(row = 2, column = arkusz.max_column).value = godzina

kolumna_powiatu = arkusz.max_column

#otworzenie pliku z powiatami
plik_z_powiatami = open('D:\\Piotr\\Magisterka\\RISE\\lista_wms_kiip.txt','r')

# iterowanie po kolejnych liniach, tworzenie 'obiektu_powiatu'
i = 2
j = 0
for linia in plik_z_powiatami:

    lista = linia.split('\t')
    obiekt_powiatu = powiat(lista[0], lista[1], lista[2], lista[5], lista[6], lista[7], lista[8], lista[9], lista[10], lista[11], lista[12], lista[13].rstrip())


    print obiekt_powiatu.nazwa, str(obiekt_powiatu.teryt)
    i += 1

    try:
        wms = WebMapService(obiekt_powiatu.adres)
    except AttributeError as at_error:      # warunek na wersje 1.3.0 wmsa np. Białystok
        if str(at_error) == "'NoneType' object has no attribute 'find'":
            wms = WebMapService(obiekt_powiatu.adres, version = '1.3.0')
    except ValueError as va_error:          # warunek na brak zdefiniowanych styli np. powiat bielski
        from owslib.testpiotr.wms import WebMapService
        wms = WebMapService(obiekt_powiatu.adres)
    except:                                 # warunek gdy nie chce się połączyc
        Raport(lista_bledow = [3])
        obiekt_powiatu.stan = 'Nie dziala'

    #jeżeli wms odpowiada, pytam o warstwy i szukam takiej, ktora ma w nazwie słowo 'działki'
    if obiekt_powiatu.stan == 'Dziala':

        X1 = float(obiekt_powiatu.X_p_losowego) - 400
        Y1 = float(obiekt_powiatu.Y_p_losowego) - 600
        X2 = float(obiekt_powiatu.X_p_losowego) + 400
        Y2 = float(obiekt_powiatu.Y_p_losowego) + 600
        bb92 = (Y1, X1, Y2, X2)

        if obiekt_powiatu.teryt == "3261": # warunek na Koszalin, ogranicznie rozmiarow
            X1 = float(obiekt_powiatu.X_p_losowego) - 100
            Y1 = float(obiekt_powiatu.Y_p_losowego) - 200
            X2 = float(obiekt_powiatu.X_p_losowego) + 100
            Y2 = float(obiekt_powiatu.Y_p_losowego) + 200
            bb92 = (Y1, X1, Y2, X2)



        if obiekt_powiatu.nazwa_warstwy == "brak":

            lista_warstw = list(wms.contents)
            for item in lista_warstw:

                """warstwa działek musi zaczynac sie od poczatku przeszukiwanego
                tekstu (symbol ^) i konczyc na koncu przeszukiwanego tekstu (symbol $).
                Może zaczynac sie z malej lub duzej litery a w srodku
                moze miec l lub ł. Jeżeli nic nie znajdzie to zwraca wartosc 'None',
                ktora pozniej konwertuje na stringa i sprawdzam czy warstwa_dzialek == 'None' """

                wyraz_szukany = re.search(u'^.zia.ki$', item, re.U)
                if str(wyraz_szukany) != 'None':
                    warstwa_dzialek = wyraz_szukany.string
                    if isinstance(warstwa_dzialek, unicode) == True:
                        warstwa_dzialek = warstwa_dzialek.encode('utf-8')
                    arkusz.cell(row = i, column = arkusz.max_column).value = warstwa_dzialek
                    obiekt_powiatu.nazwa_warstwy = warstwa_dzialek
                    break

        obiekt_powiatu.nazwa_warstwy = obiekt_powiatu.nazwa_warstwy.decode('windows-1250')
        obiekt_powiatu.nazwa_warstwy = obiekt_powiatu.nazwa_warstwy.encode('utf-8')
        ZapiszObraz(obiekt_powiatu,wms,obiekt_powiatu.nazwa_warstwy)


excel.save('raport_kiip.xlsx')
excel.close()

czas2 = time.time()
czas_sekundy = czas2 - czas1
czas_minuty = czas_sekundy / 60


print '\n'
print 'CZAS CALKOWITY (min): ' + str(round(czas_minuty,2))
print '\n' + "KONIEC"

